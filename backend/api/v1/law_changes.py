"""
LawChanges API endpoints - 법령 변경 감지 로그 조회
감지 로그 전용 (승인/반려 워크플로우 없음)
"""
import io
from typing import Optional
from datetime import datetime
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from backend.api.deps import get_db, get_current_user
from backend.models.user import User
from sqlalchemy import delete as sa_delete
from backend.models.law_change import LawChange, ApiStatus
from backend.models.law import Law
from backend.schemas.ordinance import (
    LawChangeResponse,
    LawChangeListResponse,
    LawChangeStatsResponse,
)

router = APIRouter()


def _build_law_change_response(change: LawChange) -> dict:
    """LawChange 객체를 응답 dict로 변환"""
    return {
        "id": change.id,
        "law_id": change.law_id,
        "law_name": change.law.law_name if change.law else "",
        "law_type": change.law.law_type if change.law else None,
        "revision_type": change.law.revision_type if change.law else None,
        "sync_date": change.sync_date,
        "sync_batch_id": change.sync_batch_id,
        "api_status": change.api_status.value if change.api_status else None,
        "api_message": change.api_message,
        "old_values": change.old_values,
        "new_values": change.new_values,
        "dept_name": change.dept_name,
        "dept_code": change.dept_code,
        "created_at": change.created_at,
    }


@router.get("", response_model=LawChangeListResponse)
async def get_law_changes(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    api_status: Optional[str] = None,  # success, no_response, not_found
    dept_name: Optional[str] = None,
    sync_batch_id: Optional[str] = None,
    sync_date: Optional[str] = None,  # YYYY-MM-DD 형식 날짜 필터
    search: Optional[str] = None,  # 법령명 검색
    changed_field: Optional[str] = None,  # 변경내용 필드 필터
    revision_type: Optional[str] = None,  # 제개정구분 필터
    db: AsyncSession = Depends(get_db),
):
    """
    법령 변경 감지 로그 목록 조회

    - API 상태별, 부서별, 동기화 날짜별 필터링
    - 법령명 검색, 변경내용 필드 필터링, 제개정구분 필터링
    """
    filters = []
    needs_law_join = False

    if api_status:
        try:
            api_status_enum = ApiStatus(api_status)
            filters.append(LawChange.api_status == api_status_enum)
        except ValueError:
            pass

    if dept_name:
        filters.append(LawChange.dept_name.ilike(f"%{dept_name}%"))

    if sync_batch_id:
        filters.append(LawChange.sync_batch_id == sync_batch_id)

    if sync_date:
        sync_date_obj = datetime.strptime(sync_date, "%Y-%m-%d").date()
        filters.append(func.date(LawChange.sync_date) == sync_date_obj)

    if changed_field:
        filters.append(LawChange.new_values[changed_field].isnot(None))

    if revision_type:
        needs_law_join = True
        filters.append(Law.revision_type == revision_type)

    # 총 개수 쿼리
    if search or needs_law_join:
        count_query = (
            select(func.count(LawChange.id))
            .select_from(LawChange)
            .join(Law, LawChange.law_id == Law.id)
        )
        if search:
            count_query = count_query.where(Law.law_name.ilike(f"%{search}%"))
        if filters:
            count_query = count_query.where(and_(*filters))
    else:
        count_query = select(func.count(LawChange.id)).select_from(LawChange)
        if filters:
            count_query = count_query.where(and_(*filters))

    total = await db.scalar(count_query)

    # 데이터 쿼리
    query = select(LawChange).options(selectinload(LawChange.law))
    if filters:
        query = query.where(and_(*filters))

    if search or needs_law_join:
        query = query.join(Law, LawChange.law_id == Law.id)
        if search:
            query = query.where(Law.law_name.ilike(f"%{search}%"))

    query = query.order_by(LawChange.sync_date.desc(), LawChange.id.desc())
    query = query.offset((page - 1) * size).limit(size)

    result = await db.execute(query)
    changes = result.scalars().all()

    items = [_build_law_change_response(c) for c in changes]

    return {
        "total": total or 0,
        "page": page,
        "size": size,
        "items": items,
    }


@router.get("/stats", response_model=LawChangeStatsResponse)
async def get_law_change_stats(
    sync_date: Optional[str] = None,
    sync_batch_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """법령 변경 통계 조회 (API 상태별, 부서별)"""
    date_filter = []
    if sync_batch_id:
        date_filter.append(LawChange.sync_batch_id == sync_batch_id)
    elif sync_date:
        sync_date_obj = datetime.strptime(sync_date, "%Y-%m-%d").date()
        date_filter.append(func.date(LawChange.sync_date) == sync_date_obj)

    # 전체 건수
    base_query = select(func.count(LawChange.id))
    if date_filter:
        base_query = base_query.where(and_(*date_filter))
    total = await db.scalar(base_query)

    # API 상태별 건수
    by_api_status = {}
    for status in ApiStatus:
        q = select(func.count(LawChange.id)).where(LawChange.api_status == status)
        if date_filter:
            q = q.where(and_(*date_filter))
        by_api_status[status.value] = await db.scalar(q) or 0

    # 부서별 통계
    dept_stats_query = (
        select(
            LawChange.dept_name,
            func.count(LawChange.id).label("total"),
        )
        .where(LawChange.dept_name.isnot(None))
    )
    if date_filter:
        dept_stats_query = dept_stats_query.where(and_(*date_filter))
    dept_stats_query = dept_stats_query.group_by(LawChange.dept_name).order_by(func.count(LawChange.id).desc())

    dept_result = await db.execute(dept_stats_query)
    by_dept = [
        {"dept_name": row[0], "total": row[1]}
        for row in dept_result.all()
    ]

    return {
        "total": total or 0,
        "by_api_status": by_api_status,
        "by_dept": by_dept,
    }


@router.get("/export")
async def export_law_changes(
    api_status: Optional[str] = None,
    dept_name: Optional[str] = None,
    sync_date: Optional[str] = None,
    sync_batch_id: Optional[str] = None,
    search: Optional[str] = None,
    changed_field: Optional[str] = None,
    revision_type: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """법령 변경 감지 로그 엑셀 다운로드"""
    filters = []

    if api_status:
        try:
            api_status_enum = ApiStatus(api_status)
            filters.append(LawChange.api_status == api_status_enum)
        except ValueError:
            pass

    if dept_name:
        filters.append(LawChange.dept_name.ilike(f"%{dept_name}%"))

    if sync_date:
        sync_date_obj = datetime.strptime(sync_date, "%Y-%m-%d").date()
        filters.append(func.date(LawChange.sync_date) == sync_date_obj)

    if sync_batch_id:
        filters.append(LawChange.sync_batch_id == sync_batch_id)

    if changed_field:
        filters.append(LawChange.old_values[changed_field].isnot(None))

    query = select(LawChange).options(selectinload(LawChange.law))
    if filters:
        query = query.where(and_(*filters))

    if search or revision_type:
        query = query.join(Law, LawChange.law_id == Law.id)
        if search:
            query = query.where(Law.law_name.ilike(f"%{search}%"))
        if revision_type:
            query = query.where(Law.revision_type == revision_type)

    query = query.order_by(LawChange.sync_date.desc(), LawChange.id.desc())

    result = await db.execute(query)
    changes = result.scalars().all()

    # 엑셀 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "법령변경이력"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        "법령명", "법령구분", "API상태", "소관부처",
        "변경전_공포일", "변경후_공포일", "변경전_시행일", "변경후_시행일",
        "변경전_제개정", "변경후_제개정", "동기화일시"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    api_status_map = {"success": "성공", "no_response": "응답없음", "not_found": "미발견"}

    for row_idx, change in enumerate(changes, 2):
        old_vals = change.old_values or {}
        new_vals = change.new_values or {}

        row_data = [
            change.law.law_name if change.law else "",
            change.law.law_type if change.law else "",
            api_status_map.get(change.api_status.value, change.api_status.value) if change.api_status else "",
            change.dept_name or "",
            old_vals.get("proclaimed_date", ""),
            new_vals.get("proclaimed_date", ""),
            old_vals.get("enforced_date", ""),
            new_vals.get("enforced_date", ""),
            old_vals.get("revision_type", ""),
            new_vals.get("revision_type", ""),
            change.sync_date.strftime("%Y-%m-%d %H:%M") if change.sync_date else "",
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    column_widths = [35, 12, 10, 15, 12, 12, 12, 12, 12, 12, 16]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"법령변경이력_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.get("/departments")
async def get_change_departments(
    db: AsyncSession = Depends(get_db),
):
    """변경 이력이 있는 부서 목록 조회"""
    query = (
        select(
            LawChange.dept_name,
            func.count(LawChange.id).label("total"),
        )
        .where(LawChange.dept_name.isnot(None))
        .group_by(LawChange.dept_name)
        .order_by(LawChange.dept_name)
    )
    result = await db.execute(query)
    return [
        {"dept_name": row[0], "total": row[1]}
        for row in result.all()
    ]


@router.get("/sync-batches")
async def get_sync_batches(
    db: AsyncSession = Depends(get_db),
):
    """동기화 배치 목록 조회"""
    query = (
        select(
            LawChange.sync_batch_id,
            func.min(LawChange.sync_date).label("sync_date"),
            func.count(LawChange.id).label("total"),
            func.count(LawChange.id).filter(LawChange.api_status == ApiStatus.SUCCESS).label("changed"),
            func.count(LawChange.id).filter(LawChange.api_status == ApiStatus.NO_CHANGE).label("no_change"),
            func.count(LawChange.id).filter(LawChange.api_status == ApiStatus.NO_RESPONSE).label("no_response"),
            func.count(LawChange.id).filter(LawChange.api_status == ApiStatus.NOT_FOUND).label("not_found"),
            func.count(LawChange.id).filter(LawChange.api_status == ApiStatus.ERROR).label("error"),
        )
        .where(LawChange.sync_batch_id.isnot(None))
        .group_by(LawChange.sync_batch_id)
        .order_by(func.min(LawChange.sync_date).desc())
    )
    result = await db.execute(query)
    return [
        {
            "sync_batch_id": row[0],
            "sync_date": row[1],
            "total": row[2],
            "changed": row[3],
            "no_change": row[4],
            "no_response": row[5],
            "not_found": row[6],
            "error": row[7],
        }
        for row in result.all()
    ]


@router.get("/revision-types")
async def get_revision_types(
    db: AsyncSession = Depends(get_db),
):
    """제개정구분 목록 조회 (드롭다운용)"""
    query = (
        select(Law.revision_type, func.count(LawChange.id).label("count"))
        .select_from(LawChange)
        .join(Law, LawChange.law_id == Law.id)
        .where(Law.revision_type.isnot(None))
        .group_by(Law.revision_type)
        .order_by(func.count(LawChange.id).desc())
    )
    result = await db.execute(query)
    return [
        {"revision_type": row[0], "count": row[1]}
        for row in result.all()
    ]


@router.get("/sync-dates")
async def get_sync_dates(
    db: AsyncSession = Depends(get_db),
):
    """동기화 날짜 목록 조회 (드롭다운용)"""
    query = (
        select(
            func.date(LawChange.sync_date).label("sync_date"),
            func.count(LawChange.id).label("total"),
            func.count(LawChange.id).filter(LawChange.api_status == ApiStatus.SUCCESS).label("success"),
            func.count(LawChange.id).filter(LawChange.api_status == ApiStatus.NO_RESPONSE).label("no_response"),
            func.count(LawChange.id).filter(LawChange.api_status == ApiStatus.NOT_FOUND).label("not_found"),
        )
        .group_by(func.date(LawChange.sync_date))
        .order_by(func.date(LawChange.sync_date).desc())
    )
    result = await db.execute(query)
    return [
        {
            "sync_date": str(row[0]),
            "total": row[1],
            "success": row[2],
            "no_response": row[3],
            "not_found": row[4],
        }
        for row in result.all()
    ]


@router.delete("/sync-dates/{sync_date}")
async def delete_by_sync_date(
    sync_date: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """특정 동기화 날짜의 모든 변경 기록 삭제 (관리자 전용)"""
    if current_user.user_type != "ADMIN":
        raise HTTPException(status_code=403, detail="관리자만 삭제할 수 있습니다")

    from datetime import date as date_type
    target_date = date_type.fromisoformat(sync_date)
    query = sa_delete(LawChange).where(
        func.date(LawChange.sync_date) == target_date
    )
    result = await db.execute(query)
    await db.commit()
    return {"deleted": result.rowcount}


@router.delete("/sync-batches/{batch_id}")
async def delete_by_sync_batch(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """특정 동기화 배치의 모든 변경 기록 삭제 (관리자 전용)"""
    if current_user.user_type != "ADMIN":
        raise HTTPException(status_code=403, detail="관리자만 삭제할 수 있습니다")

    query = sa_delete(LawChange).where(LawChange.sync_batch_id == batch_id)
    result = await db.execute(query)
    await db.commit()
    return {"deleted": result.rowcount}


@router.get("/{change_id}")
async def get_law_change(
    change_id: int,
    db: AsyncSession = Depends(get_db),
):
    """법령 변경 상세 조회"""
    query = (
        select(LawChange)
        .options(selectinload(LawChange.law))
        .where(LawChange.id == change_id)
    )
    result = await db.execute(query)
    change = result.scalar_one_or_none()

    if not change:
        raise HTTPException(status_code=404, detail="변경 이력을 찾을 수 없습니다.")

    return _build_law_change_response(change)


@router.get("/history/{law_id}")
async def get_law_change_history(
    law_id: int,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    """특정 법령의 변경 연혁 조회 (동기화일자 내림차순)"""
    query = (
        select(LawChange)
        .options(selectinload(LawChange.law))
        .where(LawChange.law_id == law_id)
    )

    count_query = select(func.count()).select_from(query.subquery())
    total = await db.scalar(count_query)

    query = query.order_by(LawChange.sync_date.desc(), LawChange.id.desc())
    query = query.offset((page - 1) * size).limit(size)

    result = await db.execute(query)
    changes = result.scalars().all()

    items = [_build_law_change_response(c) for c in changes]

    return {
        "total": total or 0,
        "page": page,
        "size": size,
        "items": items,
    }


@router.get("/history-summary")
async def get_law_change_history_summary(
    db: AsyncSession = Depends(get_db),
):
    """법령별 변경 연혁 요약 (총 변경 횟수, 최근 변경일)"""
    query = (
        select(
            LawChange.law_id,
            Law.law_name,
            Law.law_type,
            Law.dept_name,
            func.count(LawChange.id).label("total_changes"),
            func.max(LawChange.sync_date).label("last_sync_date"),
        )
        .join(Law, LawChange.law_id == Law.id)
        .group_by(LawChange.law_id, Law.law_name, Law.law_type, Law.dept_name)
        .order_by(func.max(LawChange.sync_date).desc())
    )
    result = await db.execute(query)

    return [
        {
            "law_id": row[0],
            "law_name": row[1],
            "law_type": row[2],
            "dept_name": row[3],
            "total_changes": row[4],
            "last_sync_date": row[5],
        }
        for row in result.all()
    ]
