"""
Ordinance API endpoints
"""
import io
from typing import List, Optional
from urllib.parse import quote
from fastapi import APIRouter, Depends, Query, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from backend.api.deps import get_db, verify_admin_password, get_current_user
from backend.models.user import User
from backend.schemas.ordinance import (
    OrdinanceResponse,
    OrdinanceListResponse,
    OrdinanceSyncRequest,
    OrdinanceSyncResponse,
    OrdinanceUploadResponse,
    OrdinanceLawMappingCreate,
    OrdinanceLawMappingUpdate,
    ParentLawCreate,
    OrdinanceCreate,
    OrdinanceCreateResponse,
    OrdinanceSearchRequest,
    OrdinanceSearchResponse,
    OrdinanceSearchResult,
    OrdinanceRegisterFromApiRequest,
    OrdinanceInfoUpdateResponse,
    OrdinanceReviewCreate,
    OrdinanceReviewUpdate,
    OrdinanceReviewResponse,
    OrdinanceReviewListResponse,
    OrdinanceReviewApprovalRequest,
    AllOrdinanceReviewsResponse,
    OrdinanceReviewWithOrdinance,
    OrdinanceDetectionResultsResponse,
    LawInfoUpdateResponse,
)
from backend.schemas.revision import DetectRequest
from backend.services.ordinance_service import OrdinanceService
from backend.services.revision_detection_service import RevisionDetectionService
from backend.core.exceptions import NotFoundError

router = APIRouter()


@router.get("", response_model=OrdinanceListResponse)
async def get_ordinances(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    category: Optional[str] = None,
    department: Optional[str] = None,
    search: Optional[str] = None,
    no_parent_law_filter: Optional[str] = None,  # "no_mapping" | "confirmed_none" | None
    needs_revision_filter: Optional[str] = None,  # "needs_revision" | "no_revision" | None
    revision_type: Optional[str] = None,  # 제개정구분 필터
    exclude_other_law_revision: bool = False,  # 타법개정 제외 여부
    review_result_filter: Optional[str] = None,  # 검토결과 필터: "개정필요" | "개정불필요" | "미검토"
    status: Optional[str] = None,  # 조례 상태 필터: "ACTIVE" | "ABOLISHED" | None(기본=ACTIVE만)
    revision_status_filter: Optional[str] = None,  # "any" | "검토대기" | "검토중" | "개정확정"
    db: AsyncSession = Depends(get_db),
):
    """Get list of ordinances"""
    service = OrdinanceService(db)
    return await service.get_list(
        page=page,
        size=size,
        category=category,
        department=department,
        search=search,
        no_parent_law_filter=no_parent_law_filter,
        needs_revision_filter=needs_revision_filter,
        revision_type=revision_type,
        exclude_other_law_revision=exclude_other_law_revision,
        review_result_filter=review_result_filter,
        status_filter=status,
        revision_status_filter=revision_status_filter,
    )


@router.get("/departments")
async def get_departments(
    db: AsyncSession = Depends(get_db),
):
    """소관부서 목록 조회 (트리용)"""
    service = OrdinanceService(db)
    return await service.get_departments()


@router.get("/revision-types")
async def get_revision_types(
    db: AsyncSession = Depends(get_db),
):
    """제개정구분 목록 조회 (드롭다운용)"""
    service = OrdinanceService(db)
    return await service.get_revision_types()


@router.post("/sync", response_model=OrdinanceSyncResponse)
async def sync_ordinances(
    request: OrdinanceSyncRequest = OrdinanceSyncRequest(),
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(verify_admin_password),
):
    """법제처 API에서 자치법규 목록을 가져와 DB에 저장 (관리자 전용)"""
    service = OrdinanceService(db)
    return await service.sync_from_moleg(org=request.org, sborg=request.sborg)


@router.post("/upload", response_model=OrdinanceUploadResponse)
async def upload_ordinances(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(verify_admin_password),
):
    """엑셀 파일로 소관부서 정보 일괄 업데이트 (관리자 전용)"""
    service = OrdinanceService(db)
    return await service.upload_from_excel(file)


@router.post("/create", response_model=OrdinanceCreateResponse)
async def create_ordinance(
    data: OrdinanceCreate,
    db: AsyncSession = Depends(get_db),
):
    """자치법규 수동 등록"""
    service = OrdinanceService(db)
    try:
        result = await service.create_ordinance(
            name=data.name,
            category=data.category,
            department=data.department,
            enacted_date=data.enacted_date,
            enforced_date=data.enforced_date,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/search-api", response_model=OrdinanceSearchResponse)
async def search_ordinance_from_api(
    request: OrdinanceSearchRequest,
):
    """
    법제처 API에서 자치법규 검색

    - 자치법규명으로 검색하여 결과 반환
    - 신규 등록 시 검색 결과에서 선택하여 등록
    """
    import httpx
    from backend.core.config import settings

    api_key = settings.MOLEG_API_KEY or "test"
    url = "http://www.law.go.kr/DRF/lawSearch.do"

    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(
                url,
                params={
                    "OC": api_key,
                    "target": "ordin",
                    "type": "JSON",
                    "query": request.query,
                    "org": request.org,
                    "sborg": request.sborg,
                    "display": 20,
                    "nw": 1,  # 현행
                },
            )

            if response.text.strip().startswith("<!DOCTYPE"):
                return OrdinanceSearchResponse(
                    success=False,
                    total=0,
                    items=[],
                    message="API 응답 오류",
                )

            response.raise_for_status()
            data = response.json()

            ordin_search = data.get("OrdinSearch", {})
            if not ordin_search:
                return OrdinanceSearchResponse(
                    success=True,
                    total=0,
                    items=[],
                    message="검색 결과가 없습니다.",
                )

            total = int(ordin_search.get("totalCnt", 0))
            items = ordin_search.get("law", [])

            if isinstance(items, dict):
                items = [items]

            results = []
            for item in items:
                results.append(OrdinanceSearchResult(
                    serial_no=item.get("자치법규일련번호", ""),
                    name=item.get("자치법규명", ""),
                    ordinance_id=item.get("자치법규ID", ""),
                    enacted_date=item.get("공포일자"),
                    promulgation_no=item.get("공포번호"),
                    revision_type=item.get("제개정구분명"),
                    org_name=item.get("지자체기관명"),
                    category=item.get("자치법규종류"),
                    enforced_date=item.get("시행일자"),
                    detail_link=item.get("자치법규상세링크"),
                    field_name=item.get("자치법규분야명"),
                ))

            return OrdinanceSearchResponse(
                success=True,
                total=total,
                items=results,
                message=f"{total}건 검색됨",
            )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"자치법규 검색 중 오류 발생: {str(e)}",
        )


@router.post("/register-from-api", response_model=OrdinanceCreateResponse)
async def register_ordinance_from_api(
    data: OrdinanceRegisterFromApiRequest,
    db: AsyncSession = Depends(get_db),
):
    """
    법제처 API 검색 결과로 자치법규 등록

    - 검색된 자치법규 정보를 기반으로 DB에 등록
    - 이미 등록된 경우 오류 반환
    """
    service = OrdinanceService(db)
    try:
        result = await service.register_from_api(
            serial_no=data.serial_no,
            name=data.name,
            ordinance_id=data.ordinance_id,
            enacted_date=data.enacted_date,
            promulgation_no=data.promulgation_no,
            revision_type=data.revision_type,
            org_name=data.org_name,
            category=data.category,
            enforced_date=data.enforced_date,
            detail_link=data.detail_link,
            field_name=data.field_name,
            department=data.department,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/update-all-info", response_model=OrdinanceInfoUpdateResponse)
async def update_all_ordinance_info(
    db: AsyncSession = Depends(get_db),
):
    """
    모든 자치법규 정보를 법제처 API로 업데이트

    - ordinances 테이블의 모든 자치법규명에 대해 법제처 API 호출
    - 공포일, 시행일, 자치법규ID 등 정보 업데이트
    """
    service = OrdinanceService(db)

    try:
        result = await service.update_all_ordinance_info()
        return OrdinanceInfoUpdateResponse(
            success=True,
            total_ordinances=result["total_ordinances"],
            updated=result["updated"],
            failed=result["failed"],
            message=f"자치법규 정보 업데이트 완료: {result['updated']}건 성공, {result['failed']}건 실패",
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"자치법규 정보 업데이트 중 오류 발생: {str(e)}",
        )


@router.get("/export")
async def export_ordinances(
    category: Optional[str] = None,
    department: Optional[str] = None,
    search: Optional[str] = None,
    no_parent_law_filter: Optional[str] = None,
    needs_revision_filter: Optional[str] = None,
    revision_type: Optional[str] = None,
    exclude_other_law_revision: bool = False,
    db: AsyncSession = Depends(get_db),
):
    """자치법규 목록 엑셀 다운로드 (필터 적용)"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime

    service = OrdinanceService(db)
    result = await service.get_list(
        page=1,
        size=10000,
        category=category,
        department=department,
        search=search,
        no_parent_law_filter=no_parent_law_filter,
        needs_revision_filter=needs_revision_filter,
        revision_type=revision_type,
        exclude_other_law_revision=exclude_other_law_revision,
    )
    items = result.get("items", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "자치법규목록"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["자치법규명", "종류", "제개정", "공포일", "시행일", "소관부서", "상위법령수", "개정여부"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, item in enumerate(items, 2):
        needs_rev = item.get("needs_revision")
        revision_text = "개정대상" if needs_rev == 1 else "대상아님" if needs_rev == 0 else "-"

        row_data = [
            item.get("name", ""),
            item.get("category", ""),
            item.get("revision_type", ""),
            str(item.get("enacted_date", "") or ""),
            str(item.get("enforced_date", "") or ""),
            item.get("department", ""),
            item.get("parent_law_count", 0),
            revision_text,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col == 8 and needs_rev == 1:
                cell.font = Font(color="FF0000")
            elif col == 8 and needs_rev == 0:
                cell.font = Font(color="00B050")

    column_widths = [40, 8, 10, 12, 12, 15, 10, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"자치법규목록_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


@router.get("/reviews-all", response_model=AllOrdinanceReviewsResponse)
async def get_all_ordinance_reviews(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    approval_status: Optional[str] = None,   # pending | approved | rejected
    review_result: Optional[str] = None,      # 개정필요 | 개정불필요 | 검토중 | 보류
    reviewer_type: Optional[str] = None,      # DEPARTMENT | GENERAL
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    전체 검토의견 목록 조회 (관리자 승인 대시보드)

    approval_status=pending 으로 필터하면 승인 대기 목록을 볼 수 있습니다.
    """
    from sqlalchemy import select, func
    from sqlalchemy.orm import selectinload
    from backend.models.ordinance_review import OrdinanceReview
    from backend.models.ordinance import Ordinance

    stmt = (
        select(OrdinanceReview)
        .options(
            selectinload(OrdinanceReview.created_by),
            selectinload(OrdinanceReview.updated_by),
            selectinload(OrdinanceReview.approved_by),
            selectinload(OrdinanceReview.ordinance),
        )
    )

    if approval_status:
        stmt = stmt.where(OrdinanceReview.approval_status == approval_status)
    if review_result:
        stmt = stmt.where(OrdinanceReview.review_result == review_result)
    if reviewer_type:
        stmt = stmt.where(OrdinanceReview.reviewer_type == reviewer_type)

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = await db.scalar(count_stmt) or 0

    stmt = stmt.order_by(OrdinanceReview.created_at.desc())
    stmt = stmt.offset((page - 1) * size).limit(size)

    result = await db.execute(stmt)
    reviews = result.scalars().all()

    items = []
    for review in reviews:
        item = OrdinanceReviewWithOrdinance.model_validate(review)
        if review.ordinance:
            item.ordinance_name = review.ordinance.name
            item.ordinance_department = review.ordinance.department
        items.append(item)

    return AllOrdinanceReviewsResponse(total=total, page=page, size=size, items=items)


@router.post("/{ordinance_id}/start-review")
async def start_ordinance_review(
    ordinance_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    검토 시작 (FR-010)
    revision_status "검토대기"→"검토중" 전환
    부서 담당자 본인 소관 조례만 접근 가능 (FR-017)
    """
    service = OrdinanceService(db)
    ordinance = await service.start_review(ordinance_id, current_user)
    return {
        "success": True,
        "revision_status": ordinance.revision_status,
        "message": "검토가 시작되었습니다.",
    }


@router.post("/{ordinance_id}/clear-revision")
async def clear_ordinance_revision(
    ordinance_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    개정확정 해제 (FR-015, 관리자 전용)
    revision_status "개정확정"→null
    """
    if current_user.user_type not in ("ADMIN", "GENERAL"):
        raise HTTPException(status_code=403, detail="관리자만 개정확정을 해제할 수 있습니다.")

    service = OrdinanceService(db)
    ordinance = await service.clear_revision_status(ordinance_id)
    return {
        "success": True,
        "revision_status": ordinance.revision_status,
        "message": "개정확정이 해제되었습니다.",
    }


@router.get("/{ordinance_id}/detection-results", response_model=OrdinanceDetectionResultsResponse)
async def get_detection_results(
    ordinance_id: int,
    db: AsyncSession = Depends(get_db),
):
    """3탭 판별 결과 조회 (캐시)"""
    service = RevisionDetectionService(db)
    try:
        return await service.get_cached_results(ordinance_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))


@router.post("/{ordinance_id}/detect", response_model=OrdinanceDetectionResultsResponse)
async def run_detection(
    ordinance_id: int,
    request: DetectRequest,
    db: AsyncSession = Depends(get_db),
):
    """3탭 판별 실행 (methods 선택 가능)"""
    service = RevisionDetectionService(db)
    try:
        selected_methods = [method.value for method in request.methods] if request.methods else None
        return await service.detect_all(ordinance_id=ordinance_id, methods=selected_methods)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"detection failed: {exc}")


@router.post("/{ordinance_id}/sync-parent-laws", response_model=LawInfoUpdateResponse)
async def sync_parent_laws(
    ordinance_id: int,
    db: AsyncSession = Depends(get_db),
):
    """해당 조례의 상위법령만 법제처 API로 동기화"""
    from sqlalchemy import select
    from backend.models.ordinance_law_mapping import OrdinanceLawMapping
    from backend.services.law_sync_service import LawSyncService

    # 상위법령 매핑에서 law PK 목록 조회
    stmt = select(OrdinanceLawMapping.law_id).where(
        OrdinanceLawMapping.ordinance_id == ordinance_id
    )
    result = await db.execute(stmt)
    law_ids = [row[0] for row in result.all()]

    if not law_ids:
        raise HTTPException(status_code=404, detail="상위법령이 없습니다.")

    service = LawSyncService(db)
    try:
        sync_result = await service.update_laws_by_ids(law_ids)
        return LawInfoUpdateResponse(
            success=True,
            total_laws=sync_result["total_laws"],
            updated=sync_result["updated"],
            failed=sync_result["failed"],
            message=f"상위법령 동기화 완료: {sync_result['updated']}건 성공, {sync_result['failed']}건 실패",
        )
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"상위법령 동기화 중 오류 발생: {str(exc)}",
        )


@router.get("/{ordinance_id}", response_model=OrdinanceResponse)
async def get_ordinance(
    ordinance_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Get ordinance by ID"""
    service = OrdinanceService(db)
    return await service.get_by_id(ordinance_id)



@router.get("/{ordinance_id}/parent-laws")
async def get_parent_laws(
    ordinance_id: int,
    db: AsyncSession = Depends(get_db),
):
    """
    조례에 매핑된 상위법령 목록 조회 (새 구조)

    Returns:
        법령 매핑 정보 목록 (법령 상세 정보 포함)
    """
    service = OrdinanceService(db)
    return await service.get_parent_laws(ordinance_id)


@router.post("/{ordinance_id}/parent-laws")
async def create_parent_law(
    ordinance_id: int,
    data: ParentLawCreate,
    db: AsyncSession = Depends(get_db),
):
    """
    상위법령 추가 (프론트엔드 호환)

    Args:
        ordinance_id: 조례 ID
        data: 상위법령 정보
    """
    service = OrdinanceService(db)
    try:
        result = await service.create_parent_law(
            ordinance_id=ordinance_id,
            law_name=data.law_name,
            law_type=data.law_type,
            proclaimed_date=data.proclaimed_date,
            enforced_date=data.enforced_date,
            related_articles=data.related_articles,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except NotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.put("/parent-laws/{parent_law_id}")
async def update_parent_law(
    parent_law_id: int,
    data: OrdinanceLawMappingUpdate,
    db: AsyncSession = Depends(get_db),
):
    """상위법령 매핑 수정 (프론트엔드 호환)"""
    service = OrdinanceService(db)
    try:
        mapping = await service.update_law_mapping(
            mapping_id=parent_law_id,
            law_name=data.law_name,
            law_type=data.law_type,
            proclaimed_date=data.proclaimed_date,
            enforced_date=data.enforced_date,
            related_articles=data.related_articles,
        )
        return {"success": True, "id": mapping.id, "message": "수정되었습니다."}
    except NotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.delete("/parent-laws/{parent_law_id}")
async def delete_parent_law(
    parent_law_id: int,
    db: AsyncSession = Depends(get_db),
):
    """상위법령 매핑 삭제 (프론트엔드 호환)"""
    service = OrdinanceService(db)
    try:
        await service.delete_law_mapping(parent_law_id)
        return {"success": True, "message": "삭제되었습니다."}
    except NotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/{ordinance_id}/no-parent-law")
async def set_no_parent_law(
    ordinance_id: int,
    db: AsyncSession = Depends(get_db),
):
    """상위법령 없음으로 설정"""
    service = OrdinanceService(db)
    try:
        result = await service.set_no_parent_law(ordinance_id, True)
        return result
    except NotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.delete("/{ordinance_id}/no-parent-law")
async def unset_no_parent_law(
    ordinance_id: int,
    db: AsyncSession = Depends(get_db),
):
    """상위법령 없음 설정 해제"""
    service = OrdinanceService(db)
    try:
        result = await service.set_no_parent_law(ordinance_id, False)
        return result
    except NotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/{ordinance_id}/law-mappings")
async def create_law_mapping(
    ordinance_id: int,
    data: OrdinanceLawMappingCreate,
    db: AsyncSession = Depends(get_db),
):
    """
    조례-법령 매핑 추가

    Args:
        ordinance_id: 조례 ID
        data: 매핑 정보 (law_id, related_articles)
    """
    service = OrdinanceService(db)
    try:
        mapping = await service.create_law_mapping(
            ordinance_id=ordinance_id,
            law_id=data.law_id,
            related_articles=data.related_articles,
        )
        return {"success": True, "id": mapping.id, "message": "매핑이 추가되었습니다."}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except NotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.put("/law-mappings/{mapping_id}")
async def update_law_mapping(
    mapping_id: int,
    data: OrdinanceLawMappingUpdate,
    db: AsyncSession = Depends(get_db),
):
    """조례-법령 매핑 수정"""
    service = OrdinanceService(db)
    try:
        mapping = await service.update_law_mapping(
            mapping_id=mapping_id,
            related_articles=data.related_articles,
        )
        return {"success": True, "id": mapping.id, "message": "수정되었습니다."}
    except NotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.delete("/law-mappings/{mapping_id}")
async def delete_law_mapping(
    mapping_id: int,
    db: AsyncSession = Depends(get_db),
):
    """조례-법령 매핑 삭제"""
    service = OrdinanceService(db)
    try:
        await service.delete_law_mapping(mapping_id)
        return {"success": True, "message": "삭제되었습니다."}
    except NotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


# ==================== 검토이력 API ====================

@router.get("/{ordinance_id}/reviews", response_model=OrdinanceReviewListResponse)
async def get_ordinance_reviews(
    ordinance_id: int,
    db: AsyncSession = Depends(get_db),
):
    """조례의 검토이력 목록 조회"""
    service = OrdinanceService(db)
    reviews = await service.get_reviews(ordinance_id)
    return OrdinanceReviewListResponse(total=len(reviews), items=reviews)


@router.post("/{ordinance_id}/reviews", response_model=OrdinanceReviewResponse)
async def create_ordinance_review(
    ordinance_id: int,
    data: OrdinanceReviewCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """조례 검토이력 추가"""
    service = OrdinanceService(db)
    try:
        review = await service.create_review(ordinance_id, data, current_user.id)
        return review
    except NotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.put("/reviews/{review_id}", response_model=OrdinanceReviewResponse)
async def update_ordinance_review(
    review_id: int,
    data: OrdinanceReviewUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """조례 검토이력 수정 (본인 작성 의견만 수정 가능)"""
    from sqlalchemy import select
    from backend.models.ordinance_review import OrdinanceReview

    # Check if review exists and user is the owner
    result = await db.execute(
        select(OrdinanceReview).where(OrdinanceReview.id == review_id)
    )
    review = result.scalar_one_or_none()
    if not review:
        raise HTTPException(status_code=404, detail="검토이력을 찾을 수 없습니다.")

    if review.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="본인이 작성한 검토이력만 수정할 수 있습니다.")

    service = OrdinanceService(db)
    try:
        review = await service.update_review(review_id, data, current_user.id)
        return review
    except NotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.delete("/reviews/{review_id}")
async def delete_ordinance_review(
    review_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """조례 검토이력 삭제 (본인 작성 의견만 삭제 가능)"""
    from sqlalchemy import select
    from backend.models.ordinance_review import OrdinanceReview

    # Check if review exists and user is the owner
    result = await db.execute(
        select(OrdinanceReview).where(OrdinanceReview.id == review_id)
    )
    review = result.scalar_one_or_none()
    if not review:
        raise HTTPException(status_code=404, detail="검토이력을 찾을 수 없습니다.")

    if review.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="본인이 작성한 검토이력만 삭제할 수 있습니다.")

    service = OrdinanceService(db)
    try:
        await service.delete_review(review_id)
        return {"success": True, "message": "삭제되었습니다."}
    except NotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/reviews/{review_id}/approve", response_model=OrdinanceReviewResponse)
async def approve_ordinance_review(
    review_id: int,
    data: OrdinanceReviewApprovalRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """검토의견 승인/반려 (관리자 전용)"""
    # 관리자 권한 확인 (ADMIN 또는 GENERAL 허용)
    if current_user.user_type not in ("ADMIN", "GENERAL"):
        raise HTTPException(status_code=403, detail="관리자만 승인/반려할 수 있습니다.")

    if data.approval_status not in ["approved", "rejected"]:
        raise HTTPException(status_code=400, detail="유효하지 않은 승인 상태입니다.")

    service = OrdinanceService(db)
    try:
        review = await service.approve_review(
            review_id=review_id,
            approval_status=data.approval_status,
            user_id=current_user.id,
            approval_note=data.approval_note,
        )
        return review
    except NotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


